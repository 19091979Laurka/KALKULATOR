"""
Moduł: Generowanie Raportów (Excel, PDF)
Tworzy raporty odszkodowawcze z wynikami analizy
"""
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False
    logger = logging.getLogger(__name__)
    logger.warning("openpyxl/pandas not available - Excel reports disabled")

logger = logging.getLogger(__name__)


def create_excel_report(
    analysis_results: List[Dict[str, Any]],
    output_path: Path,
    title: str = "Raport Analiza Roszczeń Odszkodowawczych"
) -> bool:
    """
    Tworzy raport Excel z wynikami analizy.

    Args:
        analysis_results: Lista słowników z wynikami analizy
        output_path: Ścieżka do pliku wyjściowego
        title: Tytuł raportu

    Returns:
        True jeśli sukces, False w przypadku błędu
    """

    if not HAS_EXCEL:
        logger.error("openpyxl/pandas not available")
        return False

    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Przygotuj DataFrame
        df = pd.DataFrame(analysis_results)

        # Kolumny wynikowe
        display_columns = [
            'parcel_id',
            'area_m2',
            'value_per_m2',
            'infrastructure_present',
            'voltage',
            'operator',
            'occupied_area_m2',
            'easement_pln',
            'depreciation_pln',
            'unjust_enrichment_pln',
            'total_claim_pln'
        ]

        # Wybierz tylko dostępne kolumny
        available_cols = [col for col in display_columns if col in df.columns]
        df = df[available_cols]

        # Stwórz Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(
                writer,
                sheet_name='Analiza',
                index=False,
                startrow=2
            )

            workbook = writer.book
            worksheet = writer.sheets['Analiza']

            # Tytuł
            worksheet['A1'] = title
            worksheet['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            worksheet['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            worksheet.merge_cells('A1:L1')

            # Formatowanie nagłówków
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)

            for cell in worksheet[3]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Szerokości kolumn
            column_widths = {
                'A': 18,  # parcel_id
                'B': 12,  # area_m2
                'C': 14,  # value_per_m2
                'D': 16,  # infrastructure
                'E': 12,  # voltage
                'F': 20,  # operator
                'G': 16,  # occupied_area
                'H': 16,  # easement
                'I': 16,  # depreciation
                'J': 20,  # unjust_enrichment
                'K': 16   # total_claim
            }

            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            # Formatowanie wartości pieniężnych i m²
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in worksheet.iter_rows(min_row=4, max_row=worksheet.max_row, min_col=1, max_col=len(available_cols)):
                for cell in row:
                    cell.border = border

                    # Formatowanie liczb
                    col_letter = get_column_letter(cell.column)
                    if col_letter in ['B', 'G']:  # m²
                        cell.number_format = '#,##0.00 "m²"'
                        cell.alignment = Alignment(horizontal='right')
                    elif col_letter in ['C']:  # PLN/m²
                        cell.number_format = '#,##0.00 "PLN/m²"'
                        cell.alignment = Alignment(horizontal='right')
                    elif col_letter in ['H', 'I', 'J', 'K']:  # PLN
                        cell.number_format = '#,##0.00 "PLN"'
                        cell.alignment = Alignment(horizontal='right')

            # Podsumowanie
            summary_row = worksheet.max_row + 2
            worksheet[f'A{summary_row}'] = 'PODSUMOWANIE'
            worksheet[f'A{summary_row}'].font = Font(bold=True, size=12)
            worksheet.merge_cells(f'A{summary_row}:C{summary_row}')

            # Statystyki
            stat_row = summary_row + 1
            if 'total_claim_pln' in df.columns:
                total_claim = df['total_claim_pln'].sum()
                worksheet[f'A{stat_row}'] = 'Łączna kwota roszczeń:'
                worksheet[f'C{stat_row}'] = total_claim
                worksheet[f'C{stat_row}'].number_format = '#,##0.00 "PLN"'
                worksheet[f'C{stat_row}'].font = Font(bold=True, size=12, color="C00000")

            # Meta
            meta_row = stat_row + 2
            worksheet[f'A{meta_row}'] = f"Wygenerowano: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            worksheet[f'A{meta_row}'].font = Font(italic=True, size=9)

            logger.info(f"✓ Raport Excel zapisany: {output_path}")
            return True

    except Exception as e:
        logger.error(f"Błąd tworzenia raportu Excel: {e}")
        return False


def create_summary_dict(
    analysis_results: List[Dict[str, Any]]
) -> Dict[str, Any]:
    """
    Tworzy podsumowanie statystyczne z wyników analizy.

    Returns:
        Dict z kluczowymi statystykami
    """

    try:
        total_parcels = len(analysis_results)
        with_infra = sum(1 for r in analysis_results if r.get('infrastructure_present'))

        total_claim = sum(r.get('total_claim_pln', 0) for r in analysis_results)
        total_easement = sum(r.get('easement_pln', 0) for r in analysis_results)
        total_depreciation = sum(r.get('depreciation_pln', 0) for r in analysis_results)
        total_enrichment = sum(r.get('unjust_enrichment_pln', 0) for r in analysis_results)

        avg_parcel_value = sum(
            r.get('area_m2', 0) * r.get('value_per_m2', 0)
            for r in analysis_results
        ) / max(total_parcels, 1)

        return {
            'summary': {
                'total_parcels': total_parcels,
                'parcels_with_infrastructure': with_infra,
                'parcels_percentage': round(with_infra / max(total_parcels, 1) * 100, 1),
            },
            'claims': {
                'total_claim_pln': round(total_claim, 2),
                'average_claim_pln': round(total_claim / max(with_infra, 1), 2),
                'breakdown': {
                    'easement_pln': round(total_easement, 2),
                    'depreciation_pln': round(total_depreciation, 2),
                    'unjust_enrichment_pln': round(total_enrichment, 2),
                }
            },
            'properties': {
                'average_parcel_value_pln': round(avg_parcel_value, 2),
            }
        }

    except Exception as e:
        logger.error(f"Błąd tworzenia podsumowania: {e}")
        return {}
