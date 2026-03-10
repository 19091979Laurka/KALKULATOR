#!/usr/bin/env python3
"""
Test backend kalkulatora na 99 działkach z pliku Excel.
Wczytuje kolumnę TERYT/Identyfikator Działki, dla każdej robi POST /api/analyze,
zapisuje wyniki do tests/results_99.json.
"""
import json
import os
import sys
from pathlib import Path

try:
    import openpyxl
    import httpx
except ImportError as e:
    print("Brak zależności. Zainstaluj: pip install openpyxl httpx")
    sys.exit(1)

# Ścieżki do pliku Excel (w kolejności prób)
EXCEL_PATHS = [
    os.path.expanduser("~/Downloads/Wykaz działek referencyjnych w powiecie płońskim.xlsx"),
    os.path.expanduser("~/Downloads/Wykaz działek referencyjnych w powiecie plonskim.xlsx"),
    os.path.expanduser("~/Downloads/99 dzialek.xlsx"),
]

API_URL = "http://127.0.0.1:8080/api/analyze"
RESULTS_FILE = Path(__file__).parent / "results_99.json"


def find_teryt_column(headers: list) -> int:
    """Znajdź indeks kolumny Identyfikator Działki / TERYT."""
    # Preferuj "Identyfikator Działki" (pełny ID działki) nad "Kod TERYT Gminy"
    identyfikator_idx = None
    teryt_idx = None
    for i, h in enumerate(headers):
        if not h:
            continue
        h_lower = str(h).lower()
        if "identyfikator działk" in h_lower:
            identyfikator_idx = i
        elif "teryt" in h_lower and identyfikator_idx is None:
            teryt_idx = i
    return identyfikator_idx if identyfikator_idx is not None else (teryt_idx if teryt_idx is not None else 0)


def load_parcel_ids_from_excel(excel_path: str) -> list[str]:
    """Wczytaj identyfikatory działek z pliku Excel."""
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    col_idx = find_teryt_column(headers)
    ids = []
    for row in ws.iter_rows(min_row=2):
        if col_idx < len(row) and row[col_idx].value:
            val = str(row[col_idx].value).strip()
            if val and not val.lower().startswith("identyfikator"):
                ids.append(val)
    wb.close()
    return ids


def main():
    # 1. Wczytaj Excel
    excel_path = None
    for p in EXCEL_PATHS:
        if os.path.exists(p):
            excel_path = p
            break
    if not excel_path:
        print("Nie znaleziono pliku Excel. Sprawdź ścieżki:")
        for p in EXCEL_PATHS:
            print(f"  - {p}")
        sys.exit(1)

    print(f"Wczytuję: {excel_path}")
    parcel_ids = load_parcel_ids_from_excel(excel_path)
    print(f"Wczytano {len(parcel_ids)} działek")

    # Możliwość wznowienia z poprzedniego pliku
    results = []
    done_ids = set()
    if RESULTS_FILE.exists():
        try:
            with open(RESULTS_FILE, "r", encoding="utf-8") as f:
                results = json.load(f)
            done_ids = {r.get("parcel_id") for r in results if r.get("parcel_id")}
            if done_ids:
                print(f"Wznowienie: {len(done_ids)} wyników już zapisanych")
        except Exception:
            results = []

    to_process = [p for p in parcel_ids if p not in done_ids]
    if not to_process:
        print("Wszystkie działki już przetworzone.")
        return 0

    ok_count = sum(1 for r in results if r.get("data_status") and r.get("data_status") != "ERROR")
    error_count = sum(1 for r in results if r.get("data_status") == "ERROR")
    errors_list = []
    for r in results:
        if r.get("data_status") == "ERROR":
            mr = r.get("master_record") or {}
            err = r.get("error") or mr.get("message") or ""
            errors_list.append((r["parcel_id"], str(err)[:100]))

    initial_count = len(results)
    for idx, pid in enumerate(to_process):
        i = initial_count + idx + 1
        payload = {"parcel_ids": pid}
        try:
            resp = httpx.post(API_URL, json=payload, timeout=120.0)
            if resp.status_code != 200:
                results.append({
                    "parcel_id": pid,
                    "data_status": "ERROR",
                    "error": f"HTTP {resp.status_code}: {resp.text[:200]}",
                })
                error_count += 1
                errors_list.append((pid, f"HTTP {resp.status_code}"))
                print(f"  [{i}/{len(parcel_ids)}] {pid}: ERROR (HTTP {resp.status_code})", flush=True)
            else:
                data = resp.json()
                parcels = data.get("parcels", [])
                if parcels:
                    p = parcels[0]
                    status = p.get("data_status", "UNKNOWN")
                    results.append(p)
                    if status == "ERROR":
                        error_count += 1
                        mr = p.get("master_record") or {}
                        err_msg = p.get("error") or mr.get("message") or "brak opisu"
                        errors_list.append((pid, str(err_msg)[:100]))
                        print(f"  [{i}/{len(parcel_ids)}] {pid}: ERROR - {str(err_msg)[:60]}...", flush=True)
                    else:
                        ok_count += 1
                        print(f"  [{i}/{len(parcel_ids)}] {pid}: OK ({status})", flush=True)
                else:
                    results.append({"parcel_id": pid, "data_status": "ERROR", "error": "Brak parcels w odpowiedzi"})
                    error_count += 1
                    errors_list.append((pid, "Brak parcels"))
                    print(f"  [{i}/{len(parcel_ids)}] {pid}: ERROR (brak parcels)", flush=True)
        except Exception as e:
            results.append({
                "parcel_id": pid,
                "data_status": "ERROR",
                "error": str(e),
            })
            error_count += 1
            errors_list.append((pid, str(e)[:100]))
            print(f"  [{i}/{len(parcel_ids)}] {pid}: ERROR - {e}", flush=True)

        # Zapis na bieżąco co 10 działek (odporność na przerwanie)
        if (idx + 1) % 10 == 0 or (idx + 1) == len(to_process):
            RESULTS_FILE.parent.mkdir(parents=True, exist_ok=True)
            with open(RESULTS_FILE, "w", encoding="utf-8") as f:
                json.dump(results, f, ensure_ascii=False, indent=2)

    # 3. Zapisz wyniki (pełna wersja)
    RESULTS_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(RESULTS_FILE, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)

    # 4. Podsumowanie
    print("\n" + "=" * 50)
    print("PODSUMOWANIE")
    print("=" * 50)
    print(f"OK:     {ok_count}")
    print(f"ERROR:  {error_count}")
    print(f"Łącznie: {len(results)}")
    print(f"Wyniki zapisane: {RESULTS_FILE}")
    if errors_list:
        print("\nLista błędów:")
        for pid, err in errors_list:
            print(f"  - {pid}: {err}")

    return 0 if error_count == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
